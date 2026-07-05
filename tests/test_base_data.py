from io import BytesIO

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import Workbook, load_workbook
from backend.base_data_service import normalize_stock_code, parse_base_data_workbook
from backend.database import AsyncSessionLocal, init_db
from sqlalchemy.dialects.sqlite import insert
from backend.models import Stock


def make_base_data_workbook(rows: list[list[object]] | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["一级分类", "二级分类", "三级分类", "股票名称", "股票代码"])
    for row in rows or [
        ["工业机器人", "减速器", "RV减速器", "双环传动", "002472"],
        ["工业机器人", "减速器", "RV减速器", "秦川机床", 837],
        ["工业机器人", "减速器", "谐波减速器", "速腾聚创", "02498.HK"],
        ["工业机器人", "减速器", "谐波减速器", "未上市公司", "未上市"],
        ["人形机器人", "执行系统", "丝杠/线性执行器", "拓普集团", "601689"],
        ["人形机器人", "执行系统", "丝杠/线性执行器", "五洲新春", "603667"],
    ]:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_normalize_stock_code_supports_a_hk_us_codes():
    assert normalize_stock_code(837) == "000837"
    assert normalize_stock_code("002472") == "002472"
    assert normalize_stock_code("02498.HK") == "HK02498"
    assert normalize_stock_code("HK9988") == "HK09988"
    assert normalize_stock_code("AAPL") == "US.AAPL"
    assert normalize_stock_code("US.TSLA") == "US.TSLA"
    assert normalize_stock_code("未上市") is None


def test_base_data_replace_import_overwrites_duplicate_rows():
    content = make_base_data_workbook(
        [
            ["工业机器人", "减速器", "RV减速器", "中大力德", "002896"],
            ["人形机器人", "执行系统", "旋转执行器", "中大力德", "002896"],
        ]
    )

    parsed = parse_base_data_workbook(content, overwrite_duplicates=True)

    assert parsed.skipped == []
    assert len(parsed.stocks) == 1
    assert parsed.stocks[0].industry_chain == "人形机器人"
    assert parsed.stocks[0].industry_chain_detail == "执行系统 / 旋转执行器"


def test_base_data_append_import_skips_duplicate_rows():
    content = make_base_data_workbook(
        [
            ["工业机器人", "减速器", "RV减速器", "中大力德", "002896"],
            ["人形机器人", "执行系统", "旋转执行器", "中大力德", "002896"],
        ]
    )

    parsed = parse_base_data_workbook(content)

    assert len(parsed.stocks) == 1
    assert parsed.skipped == [{"item": "中大力德(002896)", "reason": "重复股票"}]


def test_base_data_workbook_parses_hk_and_us_codes():
    content = make_base_data_workbook(
        [
            ["AI", "云服务", "大模型", "阿里巴巴", "09988.HK"],
            ["AI", "芯片", "GPU", "英伟达", "NVDA"],
        ]
    )

    parsed = parse_base_data_workbook(content)

    assert parsed.skipped == []
    assert [(item.code, item.exchange) for item in parsed.stocks] == [("HK09988", "HK"), ("US.NVDA", "US")]


@pytest.mark.asyncio
async def test_base_data_excel_import_overwrites_stock_pool():
    await init_db()
    async with AsyncSessionLocal() as session:
        stmt = insert(Stock).values(
            code="999999",
            name="旧股票",
            exchange="SH",
            industry_chain="旧分类",
            is_active=True,
        )
        await session.execute(stmt.on_conflict_do_update(index_elements=["code"], set_={"is_active": True}))
        await session.commit()

    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.post(
                "/api/base-data/stocks/import-excel",
                files={
                    "file": (
                        "base-data.xlsx",
                        make_base_data_workbook(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            stocks_response = await client.get("/api/stocks")
            disabled_response = await client.get("/api/base-data/stocks?include_inactive=true")

    payload = response.json()["data"]
    assert response.status_code == 200
    assert payload["inserted_count"] + payload["updated_count"] >= 4
    assert payload["disabled_count"] >= 1
    assert payload["skipped_count"] >= 1

    active_codes = {item["code"] for item in stocks_response.json()["data"]}
    assert {"002472", "000837", "601689", "603667"}.issubset(active_codes)
    assert "999999" not in active_codes

    all_rows = disabled_response.json()["data"]
    old_stock = next(item for item in all_rows if item["code"] == "999999")
    assert old_stock["is_active"] is False


def make_single_stock_workbook(name: str, code: str | int) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["一级分类", "二级分类", "三级分类", "股票名称", "股票代码"])
    sheet.append(["商用/特种/医疗机器人", "医疗", "手术机器人", name, code])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.asyncio
async def test_base_data_template_download_includes_current_stock_pool():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            await client.post(
                "/api/base-data/stocks/import-excel",
                files={
                    "file": (
                        "base-data.xlsx",
                        make_base_data_workbook(),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            response = await client.get("/api/base-data/stocks/template")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == ["一级分类", "二级分类", "三级分类", "股票名称", "股票代码"]
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert any(row[3] == "双环传动" and row[4] == "002472" for row in rows)


@pytest.mark.asyncio
async def test_base_data_append_import_keeps_existing_stocks_active():
    await init_db()
    async with AsyncSessionLocal() as session:
        stmt = insert(Stock).values(
            code="888888",
            name="测试存量股票",
            exchange="SH",
            industry_chain="商用/特种/医疗机器人",
            industry_chain_detail="医疗 / 手术机器人",
            is_active=True,
        )
        await session.execute(stmt.on_conflict_do_update(index_elements=["code"], set_={"is_active": True}))
        await session.commit()

    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.post(
                "/api/base-data/stocks/import-excel?mode=append",
                files={
                    "file": (
                        "append.xlsx",
                        make_single_stock_workbook("微创机器人", "688580"),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
            )
            stocks_response = await client.get("/api/base-data/stocks?include_inactive=true")

    payload = response.json()["data"]
    assert response.status_code == 200
    assert payload["disabled_count"] == 0
    rows = {item["code"]: item for item in stocks_response.json()["data"]}
    assert rows["888888"]["is_active"] is True
    assert rows["688580"]["is_active"] is True


@pytest.mark.asyncio
async def test_base_data_manual_add_upserts_and_enables_stock():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.post(
                "/api/base-data/stocks",
                json={
                    "industry_chain": "商用/特种/医疗机器人",
                    "industry_chain_detail_level2": "医疗",
                    "industry_chain_detail_level3": "手术机器人",
                    "name": "微创机器人",
                    "code": "688580",
                },
            )
            stocks_response = await client.get("/api/base-data/stocks")

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["code"] == "688580"
    assert payload["exchange"] == "SH"
    assert payload["industry_chain_detail"] == "医疗 / 手术机器人"
    assert any(item["code"] == "688580" for item in stocks_response.json()["data"])


@pytest.mark.asyncio
async def test_base_data_manual_add_rejects_unsupported_code():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.post(
                "/api/base-data/stocks",
                json={
                    "industry_chain": "商用/特种/医疗机器人",
                    "industry_chain_detail_level2": "医疗",
                    "industry_chain_detail_level3": "手术机器人",
                    "name": "未上市公司",
                    "code": "未上市",
                },
            )

    assert response.status_code == 400


@pytest.mark.asyncio
async def test_base_data_category_maintenance_crud():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            create_response = await client.post(
                "/api/base-data/categories",
                json={"industry_chain": "测试分类", "level2": "测试二级", "level3": "旋转关节"},
            )
            category_id = create_response.json()["data"]["id"]
            update_response = await client.put(
                f"/api/base-data/categories/{category_id}",
                json={"industry_chain": "测试分类", "level2": "测试二级", "level3": "灵巧手"},
            )
            list_response = await client.get("/api/base-data/categories")
            delete_response = await client.delete(f"/api/base-data/categories/{category_id}")
            list_after_delete_response = await client.get("/api/base-data/categories")

    assert create_response.status_code == 200
    assert update_response.status_code == 200
    assert update_response.json()["data"]["level3"] == "灵巧手"
    assert any(item["id"] == category_id for item in list_response.json()["data"])
    assert delete_response.status_code == 204
    assert all(item["id"] != category_id for item in list_after_delete_response.json()["data"])


@pytest.mark.asyncio
async def test_base_data_category_delete_rejects_used_category():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            categories_response = await client.get("/api/base-data/categories")
            category = categories_response.json()["data"][0]
            delete_response = await client.delete(f"/api/base-data/categories/{category['id']}")

    assert delete_response.status_code == 400


@pytest.mark.asyncio
async def test_concept_discovery_endpoint_is_removed():
    from backend.main import app

    async with app.router.lifespan_context(app):
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as client:
            response = await client.post("/api/stocks/discover-concept-sync")

    assert response.status_code in {404, 405}
