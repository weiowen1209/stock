from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from backend.models import Stock, StockCategory
from backend.schemas.base_data import BaseDataImportResult, BaseDataStockUpsert, StockCategoryWrite


REQUIRED_HEADERS = ["一级分类", "二级分类", "三级分类", "股票名称", "股票代码"]


@dataclass
class ParsedStock:
    code: str
    name: str
    exchange: str
    industry_chain: str
    industry_chain_detail: str
    core_products: str
    supply_chain_tags: str


@dataclass
class ParsedWorkbook:
    row_count: int = 0
    stocks: list[ParsedStock] = field(default_factory=list)
    skipped: list[dict[str, str]] = field(default_factory=list)


async def list_base_data_stocks(session: AsyncSession, include_inactive: bool = False) -> list[Stock]:
    stmt = select(Stock).order_by(Stock.industry_chain, Stock.industry_chain_detail, Stock.code)
    if not include_inactive:
        stmt = stmt.where(Stock.is_active.is_(True))
    result = await session.execute(stmt)
    return list(result.scalars().all())


async def list_stock_categories(session: AsyncSession) -> list[StockCategory]:
    await sync_categories_from_stocks(session)
    result = await session.execute(
        select(StockCategory).order_by(StockCategory.industry_chain, StockCategory.level2, StockCategory.level3)
    )
    return list(result.scalars().all())


async def create_stock_category(session: AsyncSession, payload: StockCategoryWrite) -> StockCategory:
    level1, level2, level3 = _normalize_category_payload(payload)
    existing = await _find_category(session, level1, level2, level3)
    if existing is not None:
        raise ValueError("分类已存在")
    row = StockCategory(industry_chain=level1, level2=level2, level3=level3)
    session.add(row)
    await session.commit()
    await session.refresh(row)
    return row


async def update_stock_category(session: AsyncSession, category_id: int, payload: StockCategoryWrite) -> StockCategory:
    row = await session.get(StockCategory, category_id)
    if row is None:
        raise ValueError("分类不存在")
    old_detail = _category_detail(row.level2, row.level3)
    level1, level2, level3 = _normalize_category_payload(payload)
    duplicate = await _find_category(session, level1, level2, level3)
    if duplicate is not None and duplicate.id != category_id:
        raise ValueError("分类已存在")
    await session.execute(
        update(Stock)
        .where(Stock.industry_chain == row.industry_chain, Stock.industry_chain_detail == old_detail)
        .values(
            industry_chain=level1,
            industry_chain_detail=_category_detail(level2, level3),
            core_products=level3,
            supply_chain_tags=";".join([level1, level2, level3]),
        )
    )
    row.industry_chain = level1
    row.level2 = level2
    row.level3 = level3
    await session.commit()
    await session.refresh(row)
    return row


async def delete_stock_category(session: AsyncSession, category_id: int) -> None:
    row = await session.get(StockCategory, category_id)
    if row is None:
        raise ValueError("分类不存在")
    usage_count = await _count_category_usage(session, row.industry_chain, row.level2, row.level3)
    if usage_count:
        raise ValueError("分类下仍有股票，不能删除")
    await session.delete(row)
    await session.commit()


async def sync_categories_from_stocks(session: AsyncSession) -> None:
    result = await session.execute(select(Stock).where(Stock.is_active.is_(True)))
    existing_result = await session.execute(select(StockCategory))
    existing = {
        (row.industry_chain, row.level2, row.level3)
        for row in existing_result.scalars().all()
    }
    created = False
    for stock in result.scalars().all():
        level2, level3 = split_industry_detail(stock.industry_chain_detail)
        if not stock.industry_chain or not level2 or not level3:
            continue
        key = (stock.industry_chain, level2, level3)
        if key in existing:
            continue
        session.add(StockCategory(industry_chain=stock.industry_chain, level2=level2, level3=level3))
        existing.add(key)
        created = True
    if created:
        await session.commit()


async def build_base_data_template(session: AsyncSession) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "股票池导入模板"
    sheet.append(REQUIRED_HEADERS)
    for stock in await list_base_data_stocks(session):
        level2, level3 = split_industry_detail(stock.industry_chain_detail)
        sheet.append([stock.industry_chain, level2, level3, stock.name, stock.code])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def import_base_data_excel(session: AsyncSession, content: bytes, mode: str = "replace") -> BaseDataImportResult:
    if mode not in {"replace", "append"}:
        raise ValueError("导入模式必须是replace或append")
    parsed = parse_base_data_workbook(content, overwrite_duplicates=mode == "replace")
    existing_result = await session.execute(select(Stock))
    existing = {stock.code: stock for stock in existing_result.scalars().all()}
    incoming_codes = {stock.code for stock in parsed.stocks}
    inserted_count = 0
    updated_count = 0
    for item in parsed.stocks:
        row = existing.get(item.code)
        if row is None:
            row = Stock(code=item.code, name=item.name, exchange=item.exchange)
            session.add(row)
            inserted_count += 1
        else:
            updated_count += 1
        row.name = item.name
        row.exchange = item.exchange
        row.industry_chain = item.industry_chain
        row.industry_chain_detail = item.industry_chain_detail
        row.core_products = item.core_products
        row.supply_chain_tags = item.supply_chain_tags
        row.is_active = True
    disabled_count = 0
    if mode == "replace":
        for code, row in existing.items():
            if code not in incoming_codes and row.is_active:
                row.is_active = False
                disabled_count += 1
    await session.commit()
    await sync_categories_from_stocks(session)
    active_result = await session.execute(select(Stock).where(Stock.is_active.is_(True)))
    active_count = len(active_result.scalars().all())
    return BaseDataImportResult(
        row_count=parsed.row_count,
        parsed_count=len(parsed.stocks),
        inserted_count=inserted_count,
        updated_count=updated_count,
        disabled_count=disabled_count,
        skipped_count=len(parsed.skipped),
        active_count=active_count,
        skipped_examples=parsed.skipped[:10],
    )


async def upsert_base_data_stock(session: AsyncSession, payload: BaseDataStockUpsert) -> Stock:
    code = normalize_stock_code(payload.code)
    if code is None:
        raise ValueError("股票代码必须是A股、港股或美股代码")
    name = _string_value(payload.name)
    level1 = _string_value(payload.industry_chain)
    level2 = _string_value(payload.industry_chain_detail_level2)
    level3 = _string_value(payload.industry_chain_detail_level3)
    if not all([name, level1, level2, level3]):
        raise ValueError("一级分类、二级分类、三级分类、股票名称和股票代码均为必填")
    if await _find_category(session, level1, level2, level3) is None:
        raise ValueError("请选择已维护的分类")
    row = await session.get(Stock, code)
    if row is None:
        row = Stock(code=code, name=name, exchange=infer_exchange(code))
        session.add(row)
    row.name = name
    row.exchange = infer_exchange(code)
    row.industry_chain = level1
    row.industry_chain_detail = f"{level2} / {level3}"
    row.core_products = level3
    row.supply_chain_tags = ";".join([level1, level2, level3])
    row.is_active = True
    await session.commit()
    await sync_categories_from_stocks(session)
    await session.refresh(row)
    return row


async def deactivate_base_data_stock(session: AsyncSession, code: str) -> None:
    normalized_code = normalize_stock_code(code)
    if normalized_code is None:
        raise ValueError("股票代码必须是6位A股代码")
    row = await session.get(Stock, normalized_code)
    if row is None or not row.is_active:
        raise ValueError("股票不存在或已删除")
    row.is_active = False
    await session.commit()


async def _find_category(session: AsyncSession, level1: str, level2: str, level3: str) -> StockCategory | None:
    result = await session.execute(
        select(StockCategory).where(
            StockCategory.industry_chain == level1,
            StockCategory.level2 == level2,
            StockCategory.level3 == level3,
        )
    )
    return result.scalar_one_or_none()


async def _count_category_usage(session: AsyncSession, level1: str, level2: str, level3: str) -> int:
    result = await session.execute(
        select(Stock).where(Stock.industry_chain == level1, Stock.industry_chain_detail == _category_detail(level2, level3))
    )
    return len(result.scalars().all())


def _normalize_category_payload(payload: StockCategoryWrite) -> tuple[str, str, str]:
    level1 = _string_value(payload.industry_chain)
    level2 = _string_value(payload.level2)
    level3 = _string_value(payload.level3)
    if not all([level1, level2, level3]):
        raise ValueError("一级分类、二级分类、三级分类均为必填")
    return level1, level2, level3


def split_industry_detail(detail: str | None) -> tuple[str, str]:
    if "/" not in (detail or ""):
        return (detail or "").strip(), ""
    level2, level3 = (detail or "").split("/", maxsplit=1)
    return level2.strip(), level3.strip()


def _category_detail(level2: str, level3: str) -> str:
    return f"{level2} / {level3}"


def parse_base_data_workbook(content: bytes, overwrite_duplicates: bool = False) -> ParsedWorkbook:
    workbook = load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]
    indexes = {header: headers.index(header) for header in REQUIRED_HEADERS if header in headers}
    missing = [header for header in REQUIRED_HEADERS if header not in indexes]
    if missing:
        raise ValueError(f"Excel缺少必需表头: {', '.join(missing)}")
    parsed = ParsedWorkbook()
    by_code: dict[str, ParsedStock] = {}
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        parsed.row_count += 1
        level1 = _string_value(row[indexes["一级分类"]])
        level2 = _string_value(row[indexes["二级分类"]])
        level3 = _string_value(row[indexes["三级分类"]])
        name = _string_value(row[indexes["股票名称"]])
        raw_code = row[indexes["股票代码"]]
        code = normalize_stock_code(raw_code)
        item = f"{name}({_string_value(raw_code)})"
        if not name or code is None:
            parsed.skipped.append({"item": item, "reason": "非A股/港股/美股代码"})
            continue
        parsed_stock = ParsedStock(
            code=code,
            name=name,
            exchange=infer_exchange(code),
            industry_chain=level1,
            industry_chain_detail=f"{level2} / {level3}" if level2 and level3 else level2 or level3,
            core_products=level3,
            supply_chain_tags=";".join([item for item in [level1, level2, level3] if item]),
        )
        if code in seen:
            if overwrite_duplicates:
                by_code[code] = parsed_stock
            else:
                parsed.skipped.append({"item": item, "reason": "重复股票"})
            continue
        seen.add(code)
        by_code[code] = parsed_stock
    parsed.stocks = list(by_code.values())
    return parsed


def normalize_stock_code(value: Any) -> str | None:
    raw = _string_value(value).upper()
    if not raw:
        return None
    if raw.endswith(".0"):
        raw = raw[:-2]
    if raw.isdigit() and len(raw) <= 6:
        raw = raw.zfill(6)
    if re.fullmatch(r"[036]\d{5}", raw):
        return raw
    hk_match = re.fullmatch(r"(?:HK)?(\d{4,5})(?:\.HK)?", raw)
    if hk_match:
        return f"HK{hk_match.group(1).zfill(5)}"
    us_match = re.fullmatch(r"(?:US\.)?([A-Z][A-Z0-9.-]{0,9})(?:\.(?:US|NASDAQ|NYSE|AMEX))?", raw)
    if us_match and not raw.endswith(".HK"):
        return f"US.{us_match.group(1)}"
    return None


def infer_exchange(code: str) -> str:
    if code.startswith("HK"):
        return "HK"
    if code.startswith("US."):
        return "US"
    if code.startswith("6"):
        return "SH"
    return "SZ"


def _string_value(value: Any) -> str:
    return str(value).strip() if value is not None else ""
